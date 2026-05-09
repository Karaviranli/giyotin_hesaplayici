import streamlit as st
import pdfplumber
import pandas as pd
from collections import Counter
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# ---------------- 1. WEB TABANLI DOSYA OKUMA ----------------
def extract_data_from_files(uploaded_files):
    all_pieces = {} 
    
    for file in uploaded_files:
        tables_to_process = []
        
        # Dosya uzantısını kontrol et
        if file.name.lower().endswith('.pdf'):
            try:
                with pdfplumber.open(file) as pdf:
                    for page in pdf.pages:
                        tables = page.extract_tables()
                        if not tables:
                            tables = page.extract_tables(table_settings={"vertical_strategy": "text", "horizontal_strategy": "text"})
                        if tables:
                            tables_to_process.extend(tables)
            except Exception as e:
                st.error(f"PDF Okuma Hatası ({file.name}): {str(e)}")

        elif file.name.lower().endswith(('.xlsx', '.xls')):
            try:
                xls = pd.ExcelFile(file)
                for sheet_name in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                    df = df.fillna("") 
                    tables_to_process.append(df.values.tolist())
            except Exception as e:
                st.error(f"Excel Okuma Hatası ({file.name}): {str(e)}")

        # Tabloları parçala
        for table in tables_to_process:
            kod_row, olcu_row, adet_row = [], [], []
            for row in table:
                clean_row = []
                for cell in row:
                    cell_str = str(cell).strip()
                    if cell_str.endswith(".0"): cell_str = cell_str[:-2]
                    clean_row.append(cell_str if cell_str else "")

                if not any(clean_row): continue
                
                first_val = ""
                data_cells = []
                for idx, val in enumerate(clean_row):
                    if val:
                        first_val = val
                        data_cells = clean_row[idx+1:]
                        break
                
                f_val_lower = first_val.lower()
                if "kod" in f_val_lower: kod_row = data_cells
                elif "ölçü" in f_val_lower or "olcu" in f_val_lower: olcu_row = data_cells
                elif "adet" in f_val_lower: adet_row = data_cells
                    
                if kod_row and olcu_row and adet_row:
                    min_len = min(len(kod_row), len(olcu_row), len(adet_row))
                    for i in range(min_len):
                        k_str = str(kod_row[i]).replace("\n", "").strip()
                        o_str = str(olcu_row[i]).replace("\n", "").strip()
                        a_str = str(adet_row[i]).replace("\n", "").strip()
                        
                        if o_str.endswith(".0"): o_str = o_str[:-2]
                        if a_str.endswith(".0"): a_str = a_str[:-2]

                        o_str = o_str.replace(".", "").replace(",", "")
                        if o_str.isdigit() and a_str.isdigit() and k_str and k_str.lower() != "none":
                            olcu_mm = int(o_str)
                            adet = int(a_str)
                            for tk_code in k_str.split("/"):
                                tk_code = tk_code.strip()
                                if tk_code:
                                    if tk_code not in all_pieces:
                                        all_pieces[tk_code] = []
                                    all_pieces[tk_code].extend([olcu_mm] * adet)
                    kod_row, olcu_row, adet_row = [], [], []
    return all_pieces

# ---------------- 2. KESİN VE MİNİMUM FİRE ALGORİTMASI ----------------
def calculate_optimal_cutting(pieces, stock_length, kerf):
    sorted_pieces = sorted(pieces, reverse=True)
    profiles = []

    for p in sorted_pieces:
        piece_with_kerf = p + kerf
        best_fit_idx = -1
        min_leftover = float('inf')

        for i, profile in enumerate(profiles):
            leftover = stock_length - (profile["used_length"] + piece_with_kerf)
            if leftover >= 0 and leftover < min_leftover:
                min_leftover = leftover
                best_fit_idx = i

        if best_fit_idx != -1:
            profiles[best_fit_idx]["pieces"].append(p)
            profiles[best_fit_idx]["used_length"] += piece_with_kerf
        else:
            profiles.append({
                "pieces": [p],
                "used_length": piece_with_kerf
            })

    result = []
    for prof in profiles:
        result.append({
            "pieces": sorted(prof["pieces"], reverse=True),
            "used_length": prof["used_length"],
            "waste": stock_length - prof["used_length"]
        })
    return result

# ---------------- 3. EXCEL OLUŞTURMA VE İNDİRME MOTORU ----------------
def create_excel_buffer(results_dict):
    wb = Workbook()
    ws = wb.active
    ws.title = "Kavira Kesim Planı"
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 4
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 18
    
    title_font = Font(color="FF0000", bold=True) 
    header_font = Font(bold=True) 
    
    current_row = 1
    col_offsets = [1, 5] 
    profile_codes = list(results_dict.keys())
    
    for i in range(0, len(profile_codes), 2):
        max_row_in_block = current_row
        
        # Sol Blok
        code1 = profile_codes[i]
        data1 = results_dict.get(code1, [])
        
        ws.cell(row=current_row, column=col_offsets[0], value=code1).font = title_font
        ws.cell(row=current_row+1, column=col_offsets[0], value="Profil No").font = header_font
        ws.cell(row=current_row+1, column=col_offsets[0]+1, value="Kesilecek Parçalar (mm)").font = header_font
        ws.cell(row=current_row+1, column=col_offsets[0]+2, value="Kalan Fire (mm)").font = header_font
        
        r_left = current_row + 2
        for idx, profile in enumerate(data1):
            pieces_str = " + ".join(map(str, profile["pieces"])) + (" mm" if profile["pieces"] else "")
            ws.cell(row=r_left, column=col_offsets[0], value=f"{idx+1}. Profil").alignment = Alignment(horizontal='left')
            ws.cell(row=r_left, column=col_offsets[0]+1, value=pieces_str).alignment = Alignment(horizontal='left')
            ws.cell(row=r_left, column=col_offsets[0]+2, value=f"{profile['waste']} mm").alignment = Alignment(horizontal='left')
            r_left += 1
            
        max_row_in_block = max(max_row_in_block, r_left)
        
        # Sağ Blok (Varsa)
        if i + 1 < len(profile_codes):
            code2 = profile_codes[i+1]
            data2 = results_dict.get(code2, [])
            
            ws.cell(row=current_row, column=col_offsets[1], value=code2).font = title_font
            ws.cell(row=current_row+1, column=col_offsets[1], value="Profil No").font = header_font
            ws.cell(row=current_row+1, column=col_offsets[1]+1, value="Kesilecek Parçalar (mm)").font = header_font
            ws.cell(row=current_row+1, column=col_offsets[1]+2, value="Kalan Fire (mm)").font = header_font
            
            r_right = current_row + 2
            for idx, profile in enumerate(data2):
                pieces_str = " + ".join(map(str, profile["pieces"])) + (" mm" if profile["pieces"] else "")
                ws.cell(row=r_right, column=col_offsets[1], value=f"{idx+1}. Profil").alignment = Alignment(horizontal='left')
                ws.cell(row=r_right, column=col_offsets[1]+1, value=pieces_str).alignment = Alignment(horizontal='left')
                ws.cell(row=r_right, column=col_offsets[1]+2, value=f"{profile['waste']} mm").alignment = Alignment(horizontal='left')
                r_right += 1
                
            max_row_in_block = max(max_row_in_block, r_right)
            
        current_row = max_row_in_block + 2 

    # Web'den indirebilmek için sanal bir hafıza (buffer) oluşturuyoruz
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ---------------- 4. KAVIRA WEB ARAYÜZÜ (STREAMLIT) ----------------
st.set_page_config(page_title="KAVIRA Kesim Motoru", page_icon="⚙️", layout="wide")

st.markdown("""
    <h1 style='text-align: center; color: #e1b12c;'>KAVIRA</h1>
    <h3 style='text-align: center; color: #7f8fa6; margin-top: -15px;'>GİYOTİN KESİM OPTİMİZASYONU</h3>
    <hr>
""", unsafe_allow_html=True)

# İki Sütunlu Tasarım
col1, col2 = st.columns([1, 2])

with col1:
    st.subheader("⚙️ Sistem Ayarları")
    stock_length = st.number_input("Profil Boyu (mm):", value=7000, step=100)
    kerf = st.number_input("Testere Firesi (mm):", value=5, step=1)
    
    st.markdown("---")
    st.subheader("📂 Sipariş Dosyaları")
    uploaded_files = st.file_uploader("PDF veya Excel yükleyin", type=['pdf', 'xlsx', 'xls'], accept_multiple_files=True)

with col2:
    if uploaded_files:
        with st.spinner("Siparişler okunuyor..."):
            loaded_data = extract_data_from_files(uploaded_files)
            
        if loaded_data:
            st.success("✅ Dosyalar başarıyla okundu!")
            
            # Veri Önizleme Tablosu
            preview_list = []
            toplam_parca = 0
            for profil_kodu in sorted(loaded_data.keys()):
                parcalar = loaded_data[profil_kodu]
                toplam_parca += len(parcalar)
                adetler = Counter(parcalar)
                for olcu, adet in sorted(adetler.items(), reverse=True):
                    preview_list.append({"Profil Kodu": profil_kodu, "Ölçü (mm)": olcu, "Adet": adet})
            
            st.markdown(f"**Toplam Kesilecek Parça:** {toplam_parca} Adet")
            st.dataframe(pd.DataFrame(preview_list), use_container_width=True, hide_index=True)
            
            # ŞALTERİ İNDİR BUTONU
            st.markdown("---")
            if st.button("⚡ ŞALTERİ İNDİR (Hesapla)", type="primary", use_container_width=True):
                with st.spinner("KAVIRA Motoru Profilleri Sıkıştırıyor..."):
                    # Optimizasyon
                    final_results = {}
                    for profil_kodu in sorted(loaded_data.keys()):
                        final_results[profil_kodu] = calculate_optimal_cutting(loaded_data[profil_kodu], stock_length, kerf)
                    
                    # Excel'i Hazırla
                    excel_buffer = create_excel_buffer(final_results)
                    
                st.success("🎉 Kesim Planı Hazır!")
                
                # Excel İndirme Butonu
                st.download_button(
                    label="📥 EXCEL'İ İNDİR",
                    data=excel_buffer,
                    file_name="Kavira_Kesim_Plani.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        else:
            st.error("❌ Dosyalardan geçerli bir alüminyum ölçüsü bulunamadı.")
    else:
        st.info("Lütfen sol taraftan hesaplanacak dosyaları yükleyin.")
